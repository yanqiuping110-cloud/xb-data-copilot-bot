"""Brief Report Excel 导出单测。"""

from io import BytesIO

from openpyxl import load_workbook

from app.brief_report.export_excel import export_turns_excel, sanitize_sheet_name, unique_sheet_names
from app.brief_report.sheet_names_llm import _fallback_sheet_names


def test_sanitize_sheet_name_removes_invalid_chars():
    assert sanitize_sheet_name("跨端/参与*人数") == "跨端参与人数"


def test_unique_sheet_names_dedupes():
    names = unique_sheet_names(["月度趋势", "月度趋势", "年级排名"])
    assert names[0] == "月度趋势"
    assert names[1] == "月度趋势(2)"
    assert names[2] == "年级排名"


def test_fallback_sheet_names_from_turns():
    turns = [
        {"question": "用图表展示本校本月跨端参与人数", "answer": "共 120 人"},
        {"question": "昨日跳绳次数", "answer": "共 80 次"},
    ]
    names = _fallback_sheet_names(turns)
    assert len(names) == 2
    assert "用图表展示" not in names[0]


def test_export_turns_excel_writes_sheets():
    turns = [
        {
            "question": "本校本月跨端参与人数",
            "columns": ["日期", "人数"],
            "rows": [["2026-07-01", 10], ["2026-07-02", 12]],
        },
        {
            "question": "年级跳绳排名",
            "columns": ["年级", "次数"],
            "rows": [["一年级", 100]],
        },
    ]
    names = ["跨端日参与", "年级跳绳排名"]
    data = export_turns_excel(turns, names)
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == names
    ws0 = wb["跨端日参与"]
    assert ws0["A1"].value.startswith("问题：")
    assert ws0["A3"].value == "日期"
    assert ws0["B4"].value == 10
