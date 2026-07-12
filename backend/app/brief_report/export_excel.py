"""问数记录 → 多 Sheet Excel 导出。"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")
_EXCEL_SHEET_MAX_LEN = 31


def sanitize_sheet_name(name: str, *, max_len: int = _EXCEL_SHEET_MAX_LEN) -> str:
    text = _INVALID_SHEET_CHARS.sub("", (name or "").strip())
    text = re.sub(r"\s+", " ", text)
    if not text:
        text = "数据表"
    if len(text) > max_len:
        text = text[: max_len - 1] + "…"
    return text


def unique_sheet_names(names: list[str]) -> list[str]:
    """去重并保证符合 Excel sheet 命名规则。"""
    used: dict[str, int] = {}
    out: list[str] = []
    for raw in names:
        base = sanitize_sheet_name(raw)
        key = base.casefold()
        count = used.get(key, 0)
        if count == 0:
            final = base
        else:
            suffix = f"({count + 1})"
            trim = _EXCEL_SHEET_MAX_LEN - len(suffix)
            final = f"{base[:trim]}{suffix}"
        used[key] = count + 1
        out.append(final)
    return out


def _autosize_columns(ws, *, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[letter].width = max(10, min(max_width, max_len + 2))


def export_turns_excel(
    turns: list[dict[str, Any]],
    sheet_names: list[str],
) -> bytes:
    """将 turns 写入 xlsx，每个 turn 一个 sheet。返回文件字节。"""
    if len(turns) != len(sheet_names):
        raise ValueError("turns 与 sheet_names 数量不一致")

    wb = Workbook()
    wb.remove(wb.active)

    for turn, sheet_name in zip(turns, sheet_names):
        ws = wb.create_sheet(title=sheet_name)
        question = (turn.get("question") or "").strip()
        if question:
            ws.append([f"问题：{question}"])
            ws["A1"].font = Font(bold=True)
            ws.append([])

        columns = turn.get("columns") or []
        rows = turn.get("rows") or []
        if columns:
            ws.append(columns)
            for row in rows:
                ws.append([_cell_value(c) for c in row])
            header_row = 3 if question else 1
            for cell in ws[header_row]:
                cell.font = Font(bold=True)
        elif rows:
            for row in rows:
                ws.append([_cell_value(c) for c in row])

        _autosize_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)
